from openpyxl import *
from datetime import datetime

wb = load_workbook('exemplo.xlsx')

sheet1 = wb["Sheet1"]

max_row = sheet1.max_row
max_column = sheet1.max_column
date_now = datetime.now()
sheet1["A8"] = date_now.strftime("%Y-%m-%d %H:%M:%S")
sheet1["B8"] = "Total"
sheet1["C8"] = "=SUM(C1:C7)"

wb.save('exemplo2.xlsx')